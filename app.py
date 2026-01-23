import os
import uuid
from flask import Flask, render_template, request, send_file, redirect, url_for
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Color
from copy import copy
import io

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

# In-memory simple cache for downloads (Not suitable for production with multiple workers)
DOWNLOAD_CACHE = {}

def workbook_to_view_data(wb):
    """
    Convert workbook to a structure for rendering in HTML.
    Returns: list of sheets, where each sheet is {'name': str, 'rows': list of lists of dicts}
    Cell dict: {'value': str, 'class': str (red/green/yellow/normal)}
    """
    sheets = []
    
    # Define color mappings based on the styles used in compare_excels
    # Red Font -> diff
    # Green Fill -> header
    # Yellow Fill -> key
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {'name': sheet_name, 'rows': []}
        
        # Determine max columns to ensure grid alignment
        # Iterate rows
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                cell_info = {'value': cell.value if cell.value is not None else ""}
                
                # Check style
                # Note: openpyxl colors can be RGB objects or legacy indexed colors. 
                # We check the specific properties we set.
                
                style_class = ""
                
                # Check Font Color (Red indicates diff)
                if cell.font and cell.font.color and cell.font.color.rgb == "FFFF0000":
                    style_class += " text-red-600 font-bold"
                
                # Check Fill Color
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color.rgb 
                    # Green: FF00FF00, Yellow: FFFFFF00, Red Fill: FFFF0000 (We used red font mainly, but fill was defined too)
                    if color == "FF00FF00":
                        style_class += " bg-green-100"
                    elif color == "FFFFFF00":
                        style_class += " bg-yellow-100"
                
                cell_info['class'] = style_class
                row_data.append(cell_info)
            sheet_data['rows'].append(row_data)
        sheets.append(sheet_data)
    return sheets

def compare_excels(file1, file2):
    wb1 = load_workbook(file1, data_only=True)
    wb2 = load_workbook(file2, data_only=True)
    
    # Create a new workbook for output
    output_wb = Workbook()
    output_wb.remove(output_wb.active) # Remove default sheet
    
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    red_font = Font(color='FFFF0000')

    # Iterate through sheets in the first workbook
    for sheet_name in wb1.sheetnames:
        ws1 = wb1[sheet_name]
        
        # Create sheet in output
        ws_out = output_wb.create_sheet(title=sheet_name)
        
        # Check if sheet exists in second workbook
        ws2 = None
        if sheet_name in wb2.sheetnames:
            ws2 = wb2[sheet_name]
            
        # Strategy: Scan rows to find "ID" column
        id_col_idx = None # 1-based index
        header_row_idx = None # 1-based index
        
        # Scan first 20 rows (or max rows) to find header
        max_scan_rows = min(20, ws1.max_row)
        for r_idx in range(1, max_scan_rows + 1):
            row = ws1[r_idx]
            for cell in row:
                if cell.value:
                    val_str = str(cell.value).lower()
                    if "id" in val_str or "sku" in val_str or "#" in val_str:
                        id_col_idx = cell.column
                        header_row_idx = cell.row
                        break
            if id_col_idx:
                break
        
        # If ID column found and ws2 exists, build index for ws2
        ws2_index = {} # Map ID -> Row Object (or Row Index)
        if id_col_idx and ws2 and header_row_idx:
            for row in ws2.iter_rows(min_row=header_row_idx + 1):
                try:
                    if len(row) >= id_col_idx:
                        id_cell = row[id_col_idx - 1]
                        val = id_cell.value
                        if val is not None:
                            ws2_index[val] = row
                except IndexError:
                    pass

        # Iterate through rows and columns of File 1
        for row in ws1.iter_rows():
            current_row_idx = row[0].row
            
            # Find matching row in File 2
            row2 = None
            if ws2:
                # Only try to match data rows (rows after the header)
                if header_row_idx and current_row_idx > header_row_idx:
                     if id_col_idx and ws2_index:
                        # Look up by ID
                        try:
                             if len(row) >= id_col_idx:
                                id_val = row[id_col_idx - 1].value
                                if id_val in ws2_index:
                                    row2 = ws2_index[id_val]
                        except:
                            pass
                     else:
                        # Fallback to positional match
                         try:
                            row2 = ws2[current_row_idx]
                         except IndexError:
                            row2 = None
                else:
                    # For header row or pre-header rows, try to match by position
                    if header_row_idx and current_row_idx == header_row_idx:
                         try:
                             row2 = ws2[current_row_idx]
                         except IndexError:
                             row2 = None
                    else:
                        # Pre-header rows
                        try:
                             row2 = ws2[current_row_idx]
                        except IndexError:
                             row2 = None


            for i, cell in enumerate(row):
                # Copy value to output
                new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Apply styles
                if header_row_idx and cell.row == header_row_idx:
                    new_cell.fill = green_fill
                elif id_col_idx and cell.column == id_col_idx:
                     # Only highlight ID column in data rows
                    if header_row_idx and cell.row > header_row_idx:
                        new_cell.fill = yellow_fill
                
                # Compare
                is_diff = False
                if ws2:
                    if row2:
                        try:
                            # Verify row2 has this column
                            if i < len(row2):
                                cell2 = row2[i]
                                if cell.value != cell2.value:
                                    is_diff = True
                            else:
                                is_diff = True # Column missing in row2
                        except:
                            is_diff = True
                    else:
                        is_diff = True # Row not found in File 2
                else:
                     is_diff = True # Sheet doesn't exist in file 2
                
                if is_diff:
                    # Make value red if different
                    new_cell.font = red_font

    return output_wb

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/compare', methods=['POST'])
def compare():
    if 'file1' not in request.files or 'file2' not in request.files:
        return 'No files uploaded', 400
    
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    if file1.filename == '' or file2.filename == '':
        return 'No selected file', 400

    if file1 and file2:
        output_wb = compare_excels(file1, file2)
        
        # 1. Generate Table Project View Data
        view_data = workbook_to_view_data(output_wb)
        
        # 2. Save for download
        download_id = str(uuid.uuid4())
        output_stream = io.BytesIO()
        output_wb.save(output_stream)
        output_stream.seek(0)
        DOWNLOAD_CACHE[download_id] = output_stream
        
        return render_template('index.html', result=view_data, download_id=download_id)
    
    return redirect(url_for('index'))

@app.route('/download/<download_id>')
def download_file(download_id):
    if download_id in DOWNLOAD_CACHE:
        stream = DOWNLOAD_CACHE[download_id]
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name='comparison_result.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return "File not found or expired", 404

if __name__ == '__main__':
    app.run(debug=True, port=5000)
