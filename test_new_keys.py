
import os
import unittest
from openpyxl import Workbook
from app import compare_excels

class TestKeyCompare(unittest.TestCase):
    def setUp(self):
        # Create dummy Excel files
        self.file1 = 'test_file1.xlsx'
        self.file2 = 'test_file2.xlsx'

    def tearDown(self):
        # Clean up
        if os.path.exists(self.file1):
            os.remove(self.file1)
        if os.path.exists(self.file2):
            os.remove(self.file2)

    def create_excel(self, filename, data, header):
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in data:
            ws.append(row)
        wb.save(filename)

    def test_sku_key(self):
        # File 1: SKU is key
        header = ['Name', 'SKU', 'Price']
        data1 = [
            ['Apple', 'A100', 10],
            ['Banana', 'B200', 20]
        ]
        self.create_excel(self.file1, data1, header)

        # File 2: Same SKUs, different order
        data2 = [
            ['Banana', 'B200', 20],
            ['Apple', 'A100', 12] # Price changed
        ]
        self.create_excel(self.file2, data2, header)
        
        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # Row 2 (Apple) should have red font in Price col (col 3)
        # Note: In output, rows follow File 1 order.
        # Row 2 is Apple.
        
        price_cell = ws_out.cell(row=2, column=3)
        self.assertIsNotNone(price_cell.font.color, "Font color should be set")
        self.assertEqual(price_cell.font.color.rgb, "FFFF0000", "Price should be red")
        
        name_cell = ws_out.cell(row=2, column=1)
        # Name should NOT be red
        # openpyxl might return a Color object with rgb=None for default theme colors
        if name_cell.font.color:
             self.assertNotEqual(name_cell.font.color.rgb, "FFFF0000", "Name should not be red")

    def test_hash_key(self):
        # File 1: # is key
        header = ['#', 'Item']
        data1 = [
            [1, 'One'],
            [2, 'Two']
        ]
        self.create_excel(self.file1, data1, header)

        # File 2
        data2 = [
            [2, 'Two'],
            [1, 'OneDiff'] # Changed
        ]
        self.create_excel(self.file2, data2, header)

        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # Row 2 (ID 1) should have diff in Item (col 2)
        item_cell = ws_out.cell(row=2, column=2)
        self.assertEqual(item_cell.font.color.rgb, "FFFF0000")

if __name__ == '__main__':
    unittest.main()
