
import os
import unittest
from openpyxl import Workbook
from app import compare_excels

class TestDynamicHeader(unittest.TestCase):
    def setUp(self):
        self.file1 = 'test_dyn_1.xlsx'
        self.file2 = 'test_dyn_2.xlsx'

    def tearDown(self):
        if os.path.exists(self.file1):
            os.remove(self.file1)
        if os.path.exists(self.file2):
            os.remove(self.file2)

    def create_excel(self, filename, data, header, pre_header_rows=0):
        wb = Workbook()
        ws = wb.active
        
        # Add pre-header rows
        for _ in range(pre_header_rows):
            ws.append(['Title or Empty'])
            
        ws.append(header)
        for row in data:
            ws.append(row)
        wb.save(filename)

    def test_delayed_header(self):
        # Header is on Row 3
        pre_rows = 2
        header = ['Name', 'ID', 'Value']
        data1 = [['A', 1, 100], ['B', 2, 200]]
        # File 2: Same data
        data2 = [['A', 1, 100], ['B', 2, 200]]
        
        self.create_excel(self.file1, data1, header, pre_rows)
        self.create_excel(self.file2, data2, header, pre_rows)
        
        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # Row 3 should be Green (Header)
        # Note: In output, empty/title rows are copied.
        # Row 1: Title
        # Row 2: Title
        # Row 3: Header
        
        header_cell = ws_out.cell(row=3, column=2)
        self.assertEqual(header_cell.fill.start_color.rgb, "FF00FF00", "Row 3 should be green (Header)")
        
        # Row 4 (Data) Col 2 (ID) should be yellow
        id_cell = ws_out.cell(row=4, column=2)
        self.assertEqual(id_cell.fill.start_color.rgb, "FFFFFF00", "Row 4 Col 2 should be yellow (ID Data)")
        
        # Row 1/2 Col 2 should NOT be yellow (it's pre-header)
        pre_cell = ws_out.cell(row=2, column=2)
        # Check it's not yellow
        if pre_cell.fill and pre_cell.fill.start_color.type == 'rgb':
             self.assertNotEqual(pre_cell.fill.start_color.rgb, "FFFFFF00")

    def test_diff_with_offset_header(self):
        # Header on Row 2
        pre_rows = 1
        header = ['Item', 'SKU', 'Q']
        data1 = [['A', 'S1', 10]]
        data2 = [['A', 'S1', 99]] # Diff in Q
        
        self.create_excel(self.file1, data1, header, pre_rows)
        self.create_excel(self.file2, data2, header, pre_rows)
        
        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # Row 2 = Header (Green)
        self.assertEqual(ws_out.cell(row=2, column=2).fill.start_color.rgb, "FF00FF00")
        
        # Row 3 = Data (SKU=S1). Col 3 (Q) should be Red (Diff)
        diff_cell = ws_out.cell(row=3, column=3)
        self.assertEqual(diff_cell.font.color.rgb, "FFFF0000", "Diff value should be red")


if __name__ == '__main__':
    unittest.main()
