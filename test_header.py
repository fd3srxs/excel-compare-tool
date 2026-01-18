
import os
import unittest
from openpyxl import Workbook
from app import compare_excels

class TestHeaderHighlight(unittest.TestCase):
    def setUp(self):
        self.file1 = 'test_header_1.xlsx'
        self.file2 = 'test_header_2.xlsx'

    def tearDown(self):
        if os.path.exists(self.file1):
            os.remove(self.file1)
        if os.path.exists(self.file2):
            os.remove(self.file2)

    def create_excel(self, filename, data, header):
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in data:
            ws.append(row)
        wb.save(filename)

    def test_green_header(self):
        # Header has ID -> should be key
        header = ['Name', 'ID', 'Value']
        data1 = [['A', 1, 100], ['B', 2, 200]]
        data2 = [['A', 1, 100], ['B', 2, 200]]
        
        self.create_excel(self.file1, data1, header)
        self.create_excel(self.file2, data2, header)
        
        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # Check Header (Row 1)
        # All columns in header should be green
        for col in range(1, 4):
            cell = ws_out.cell(row=1, column=col)
            self.assertEqual(cell.fill.start_color.rgb, "FF00FF00", f"Header cell {col} should be green")
            
        # Check Data (Row 2)
        # Col 2 (ID) should be yellow (from previous rule)
        id_cell = ws_out.cell(row=2, column=2)
        self.assertEqual(id_cell.fill.start_color.rgb, "FFFFFF00", "Data ID column should be yellow")
        
        # Col 1 (Name) should NOT be green
        name_cell = ws_out.cell(row=2, column=1)
        if name_cell.fill and name_cell.fill.start_color.type == 'rgb':
             self.assertNotEqual(name_cell.fill.start_color.rgb, "FF00FF00")

if __name__ == '__main__':
    unittest.main()
