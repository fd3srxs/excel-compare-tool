import unittest
import os
from openpyxl import Workbook, load_workbook
from app import compare_excels

class TestExcelCompare(unittest.TestCase):
    def setUp(self):
        # Create dummy file 1
        self.wb1 = Workbook()
        ws1 = self.wb1.active
        ws1['A1'] = 10
        ws1['B1'] = "Hello"
        ws1['C1'] = 3.14
        self.wb1.save('test1.xlsx')
        
        # Create dummy file 2
        self.wb2 = Workbook()
        ws2 = self.wb2.active
        ws2['A1'] = 10     # Same
        ws2['B1'] = "World" # Diff
        ws2['C1'] = 3.14   # Same
        self.wb2.save('test2.xlsx')

    def tearDown(self):
        if os.path.exists('test1.xlsx'):
            os.remove('test1.xlsx')
        if os.path.exists('test2.xlsx'):
            os.remove('test2.xlsx')
        if os.path.exists('test_output.xlsx'):
            os.remove('test_output.xlsx')

    def test_compare(self):
        output_wb = compare_excels('test1.xlsx', 'test2.xlsx')
        output_wb.save('test_output.xlsx')
        
        wb_out = load_workbook('test_output.xlsx')
        ws_out = wb_out.active
        
        # Check A1: Same (No red connection)
        # Note: openpyxl returns color as ARGB hex string. Red is usually FFFF0000
        # If no color is set, it might be None or theme connection.
        
        # Check B1: Different (Should be red)
        font_b1 = ws_out['B1'].font
        print(f"B1 Font Color: {font_b1.color.rgb}")
        self.assertEqual(font_b1.color.rgb, 'FFFF0000')

        # Check A1: Same (Should not be red)
        # Default font color is usually None (automatic) or '00000000' (black)
        font_a1 = ws_out['A1'].font
        print(f"A1 Font Color: {font_a1.color.rgb}")
        self.assertNotEqual(font_a1.color.rgb, 'FFFF0000')

if __name__ == '__main__':
    unittest.main()
