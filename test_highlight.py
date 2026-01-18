
import os
import unittest
from openpyxl import Workbook
from app import compare_excels

class TestKeyHighlight(unittest.TestCase):
    def setUp(self):
        self.file1 = 'test_highlight_1.xlsx'
        self.file2 = 'test_highlight_2.xlsx'

    def tearDown(self):
        if os.path.exists(self.file1):
            os.remove(self.file1)
        if os.path.exists(self.file2):
            os.remove(self.file2)

    def create_excel(self, filename, data, header):
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in data:
            ws.append(row)
        wb.save(filename)

    def test_yellow_highlight(self):
        # Header has ID -> should be key
        header = ['Name', 'ID', 'Value']
        data1 = [['A', 1, 100], ['B', 2, 200]]
        data2 = [['A', 1, 100], ['B', 2, 200]]
        
        self.create_excel(self.file1, data1, header)
        self.create_excel(self.file2, data2, header)
        
        output_wb = compare_excels(self.file1, self.file2)
        ws_out = output_wb.active
        
        # ID is column 2
        # Check Row 1 (Header) Col 2
        header_cell = ws_out.cell(row=1, column=2)
        self.assertEqual(header_cell.fill.start_color.rgb, "FFFFFF00", "Header ID column should be yellow")
        
        # Check Row 2 (Data) Col 2
        data_cell = ws_out.cell(row=2, column=2)
        self.assertEqual(data_cell.fill.start_color.rgb, "FFFFFF00", "Data ID column should be yellow")
        
        # Check Row 2 Col 1 (Name) -> Should not be yellow
        name_cell = ws_out.cell(row=2, column=1)
        # Default fill is usually PatternFill() with no color or different type. 
        # But explicitly checks it's NOT yellow.
        if name_cell.fill.start_color.type == 'rgb':
             self.assertNotEqual(name_cell.fill.start_color.rgb, "FFFFFF00")

if __name__ == '__main__':
    unittest.main()
