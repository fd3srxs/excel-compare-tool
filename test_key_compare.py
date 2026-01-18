import unittest
import os
from openpyxl import Workbook, load_workbook
from app import compare_excels

class TestKeyCompare(unittest.TestCase):
    def setUp(self):
        # Case 1: Disordered rows, same content
        self.wb1 = Workbook()
        ws1 = self.wb1.active
        ws1.append(["ID", "Name", "Value"])
        ws1.append([1, "Alice", 100])
        ws1.append([2, "Bob", 200])
        self.wb1.save('test_key_1.xlsx')
        
        self.wb2 = Workbook()
        ws2 = self.wb2.active
        ws2.append(["ID", "Name", "Value"])
        ws2.append([2, "Bob", 200]) # Row 2 in file 2 corresponds to Row 3 in file 1
        ws2.append([1, "Alice", 100]) # Row 3 in file 2 corresponds to Row 2 in file 1
        self.wb2.save('test_key_2.xlsx')

        # Case 2: Disordered rows, different content
        self.wb3 = Workbook() # Base
        ws3 = self.wb3.active
        ws3.append(["ID", "Name", "Value"])
        ws3.append([1, "Alice", 100])
        ws3.append([2, "Bob", 200])
        self.wb3.save('test_diff_1.xlsx')

        self.wb4 = Workbook() # Diff
        ws4 = self.wb4.active
        ws4.append(["ID", "Name", "Value"])
        ws4.append([2, "Bob", 200])
        ws4.append([1, "Alice", 999]) # Value different for ID 1
        self.wb4.save('test_diff_2.xlsx')

    def tearDown(self):
        files = ['test_key_1.xlsx', 'test_key_2.xlsx', 'test_key_out.xlsx',
                 'test_diff_1.xlsx', 'test_diff_2.xlsx', 'test_diff_out.xlsx']
        for f in files:
            if os.path.exists(f):
                os.remove(f)

    def test_reorder_match(self):
        """Test that reordered rows with same ID content are found and NOT marked red"""
        output_wb = compare_excels('test_key_1.xlsx', 'test_key_2.xlsx')
        output_wb.save('test_key_out.xlsx')
        
        wb_out = load_workbook('test_key_out.xlsx')
        ws_out = wb_out.active
        
        # ID 1 is at Row 2. Should match ID 1 in File 2 (which is at Row 3).
        # Values are same, so no red font.
        cell_val = ws_out.cell(row=2, column=3) # Value 100
        print(f"Test 1 (Match) - Cell Value: {cell_val.value}, Font: {cell_val.font.color.rgb if cell_val.font.color else 'None'}")
        
        # We expect NO red color
        if cell_val.font.color:
             self.assertNotEqual(cell_val.font.color.rgb, 'FFFF0000')

    def test_reorder_diff(self):
        """Test that reordered rows with different content are found and marked red"""
        output_wb = compare_excels('test_diff_1.xlsx', 'test_diff_2.xlsx')
        output_wb.save('test_diff_out.xlsx')
        
        wb_out = load_workbook('test_diff_out.xlsx')
        ws_out = wb_out.active
        
        # ID 1 is at Row 2 in straight file.
        # In File 2, ID 1 has Value 999.
        # So Cell (Row 2, Column 3) should be 100 (from File 1) but marked RED because Diff.
        
        cell_val = ws_out.cell(row=2, column=3)
        print(f"Test 2 (Diff) - Cell Value: {cell_val.value}, Font: {cell_val.font.color.rgb if cell_val.font.color else 'None'}")
        
        self.assertEqual(cell_val.font.color.rgb, 'FFFF0000')

if __name__ == '__main__':
    unittest.main()
